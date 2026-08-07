# -*- coding: utf-8 -*-
"""generate_labels_v3.py - Gear fatigue paper v4.8/4.9 -> output/labels.json
Reads current output/*.json and builds:
  - output/labels.json            (paper_check Layer-1 label registry)
  - <numbers ledger>.csv / .xlsx  (every key number + source, v4.8 baseline)
Values are pulled programmatically from the JSONs so they stay current.
"""
import json, os, math, re

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "output")

def load(name):
    with open(os.path.join(OUT, name), encoding="utf-8") as f:
        return json.load(f)

anova    = load("aft_anova.json")
diag     = load("aft_diagnostics.json")
summary  = load("sweep_summary.json")
weibull  = load("weibull_aft.json")
multiaft = load("multi_aft_comparison.json")
sobol    = load("sobol_vs_aft.json")
nbp      = load("native_bias_propagation.json")
nss      = load("native_stress_sensitivity.json")
threshold= load("threshold_sensitivity.json")
ka       = load("ka_sensitivity.json")
ysa      = load("ysa_sensitivity.json")
torque   = load("multi_torque_aft.json")
sim      = load("simulation_validation.json")
twolevel = load("two_level_variable_amplitude.json")
gradient = load("sobol_censoring_gradient.json")
wang     = load("benchmark_wang2023.json")
msn      = load("multi_sn_sensitivity.json")

labels = {}
rows = []

def add(key, value, note, section, src_file, src_key, match="AUTO"):
    m = re.search(r"([A-Za-z0-9_\-]+\.json)", src_file)
    si = m.group(1) if m else ""
    labels[key] = {"value": value, "note": note,
                   "source_script": src_file, "source_input": si}
    rows.append([section, note, value, src_file, src_key, match])

def r(x, nd=2):
    if isinstance(x, float) and math.isfinite(x):
        return round(x, nd)
    return x

# ---------- design / sweep ----------
add("n_combinations", summary["n_combinations"], "Total factorial combinations (5 factors x 2-4 levels)", "Abstract / Table 2", "sweep_summary.json", "n_combinations")
add("n_runouts", summary["n_runouts"], "Run-out (censored) entries", "Abstract / Sec 3.6", "sweep_summary.json", "n_runouts")
add("n_valid", summary["n_valid"], "Finite-life (observed) entries", "Abstract / Sec 3.6", "sweep_summary.json", "n_valid")
add("n_low_cycle", summary["n_low_cycle"], "Low-cycle entries (<10^4)", "Sec 3.6", "sweep_summary.json", "n_low_cycle")
add("min_log10_Nf", r(summary["min_log10_Nf"]), "Minimum log10(Nf)", "Sec 3.6 / Conclusions", "sweep_summary.json", "min_log10_Nf")
add("max_log10_Nf", r(summary["max_log10_Nf"]), "Maximum log10(Nf)", "Sec 3.6 / Conclusions", "sweep_summary.json", "max_log10_Nf")
add("spread_dex", r(summary["spread_dex"]), "log10(Nf) spread (dex)", "Abstract / Sec 3.6 / Conclusions", "sweep_summary.json", "spread_dex")
add("module_mn", summary["gear_module_mm"], "Normal module (mm)", "Table 1 (tab:gear)", "sweep_summary.json", "gear_module_mm")
add("teeth_pinion", summary["teeth"], "Pinion teeth z1", "Table 1 (tab:gear)", "sweep_summary.json", "teeth")
add("material", summary["material"], "Material", "Table 1 (tab:gear)", "sweep_summary.json", "material")
add("torque_Nm", summary["torque_Nm"], "Reference input torque (Nm)", "Table 1 / Abstract", "sweep_summary.json", "torque_Nm")

# ---------- AFT ANOVA ----------
for row in anova["anova_table"]:
    fac = row["factor"]
    key = fac.replace(" ", "_").replace("x_", "x_")
    add(key + "_var_pct", row["variance_frac_%"], fac + " variance fraction (%)", "Table 2 (tab:anova)", "aft_anova.json", "anova_table[" + fac + "].variance_frac_%")
    add(key + "_dloglik", r(row["delta_logLik"]), fac + " delta log-likelihood", "Table 2 (tab:anova)", "aft_anova.json", "anova_table[" + fac + "].delta_logLik")
add("n_censored", anova["n_censored"], "Censored (run-out) count in AFT", "Sec 3.1 / Conclusions", "aft_anova.json", "n_censored")
add("n_observed", anova["n_observed"], "Observed (finite) count in AFT", "Sec 3.1 / Conclusions", "aft_anova.json", "n_observed")
add("sigma_aft", anova["sigma_aft"], "AFT log-normal sigma", "Sec 3.2", "aft_anova.json", "sigma_aft")
add("logLik_full", anova["logLik_full"], "Full-model log-likelihood", "Sec 3.2", "aft_anova.json", "logLik_full")
add("bootstrap_n", anova["bootstrap"]["n_bootstrap"], "Bootstrap draws", "Abstract / Sec 3.8 / Data Availability", "aft_anova.json", "bootstrap.n_bootstrap")
add("bootstrap_seed", anova["bootstrap"]["rng_seed"], "Bootstrap RNG seed", "Data Availability", "aft_anova.json", "bootstrap.rng_seed")
for fac, d in anova["bootstrap"]["noise_floor"].items():
    key = fac.replace(" ", "_")
    add(key + "_boot_mean", r(d["mean"]), fac + " bootstrap mean (%)", "Sec 3.8 / Fig 5", "aft_anova.json", "bootstrap.noise_floor." + fac + ".mean")
    add(key + "_boot_sn", d["s_n"], fac + " bootstrap S/N ratio", "Fig 5", "aft_anova.json", "bootstrap.noise_floor." + fac + ".s_n")

# ---------- diagnostics ----------
add("shapiro_W", r(diag["shapiro_wilk"]["statistic"]), "Shapiro-Wilk W (observed residuals)", "Sec 3.3 / Diagnostics", "aft_diagnostics.json", "shapiro_wilk.statistic")
add("shapiro_p", diag["shapiro_wilk"]["p_value"], "Shapiro-Wilk p-value", "Sec 3.3 / Diagnostics", "aft_diagnostics.json", "shapiro_wilk.p_value")
add("qq_R2", r(diag["qq_plot_r_squared"]), "QQ-plot R^2", "Sec 3.3 / Fig (diag_qq)", "aft_diagnostics.json", "qq_plot_r_squared")
add("cox_snell_mean", diag["cox_snell"]["mean"], "Cox-Snell residual mean", "Sec 3.3", "aft_diagnostics.json", "cox_snell.mean")
add("cox_snell_KS", r(diag["cox_snell"]["ks_statistic_vs_exp1"]), "Cox-Snell KS vs Exp(1)", "Sec 3.3", "aft_diagnostics.json", "cox_snell.ks_statistic_vs_exp1")
add("martingale_mean", diag["martingale"]["mean"], "Martingale residual mean", "Sec 3.3", "aft_diagnostics.json", "martingale.mean")
add("martingale_range", [r(diag["martingale"]["range"][0]), r(diag["martingale"]["range"][1])], "Martingale residual range", "Sec 3.3", "aft_diagnostics.json", "martingale.range")
for fac in ["mean_stress_method", "size_surface_standard", "rz_level", "sn_source", "damage_method"]:
    lv = diag["levene_test"][fac]
    add("levene_" + fac + "_p", lv["p_value"], "Levene p (" + fac + ")", "Sec 3.3", "aft_diagnostics.json", "levene_test." + fac + ".p_value")

# ---------- distribution comparison ----------
add("weibull_deltaAIC", weibull["delta_AIC_weibull_minus_lognormal"], "Weibull-AFT vs log-normal delta AIC", "Sec 3.9 / Conclusions", "weibull_aft.json", "delta_AIC_weibull_minus_lognormal")
add("lognormal_AIC", multiaft["distributions"]["log_normal"]["AIC"], "Log-normal AFT AIC", "Sec 3.3", "multi_aft_comparison.json", "distributions.log_normal.AIC")
add("loglogistic_AIC", multiaft["distributions"]["log_logistic"]["AIC"], "Log-logistic AFT AIC", "Sec 3.3", "multi_aft_comparison.json", "distributions.log_logistic.AIC")
add("gamma_AIC", multiaft["distributions"]["gamma"]["AIC"], "Gamma AFT AIC", "Sec 3.3", "multi_aft_comparison.json", "distributions.gamma.AIC")

# ---------- Sobol vs AFT ----------
for c in sobol["comparison"]:
    add("sobol_delta_" + c["factor"].replace(" ", "_"), c["delta_S1_minus_AFT_pp"],
        "Sobol' S1 minus AFT (pp): " + c["factor"], "Table 4 (tab:aft_vs_anova)", "sobol_vs_aft.json", "comparison[" + c["factor"] + "].delta_S1_minus_AFT_pp")

# ---------- Y_Sa sensitivity (Fig 1) ----------
base_std = ysa["baseline"]["variance_fractions"]["size_surface_standard"]
devs = {}
for f in ["0.95", "1.0", "1.05", "1.1", "1.15"]:
    devs[f] = round(base_std - ysa["sensitivity"][f]["variance_fractions"]["size_surface_standard"], 1)
max_dev = max(abs(v) for v in devs.values())
add("ysa_baseline_std_pct", base_std, "Standard variance at baseline Y_Sa=1.55 (%)", "Fig 1 caption / Sec 2.3", "ysa_sensitivity.json", "baseline.variance_fractions.size_surface_standard")
add("ysa_max_dev_pp", max_dev, "Max |deviation| of standard share over 0.95-1.15x perturbation (pp)", "Fig 1 caption", "ysa_sensitivity.json", "sensitivity[0.95..1.15]", "COMPUTED")
print("Y_Sa deviations over 0.95-1.15x:", devs)

# ---------- native stress chain ----------
n_base = nss["baseline"]["variance_fractions"]["size_surface_standard"]
n_nat  = nss["combined_native"]["variance_fractions"]["size_surface_standard"]
add("native_baseline_std_pct", n_base, "Standard share, 108-combo subset, uniform Y_Sa (%)", "Sec 2.3 / Sec 4.6", "native_stress_sensitivity.json", "baseline.variance_fractions.size_surface_standard")
add("native_combined_std_pct", n_nat, "Standard share with native J/Kf (%)", "Sec 2.3 / Sec 4.6", "native_stress_sensitivity.json", "combined_native.variance_fractions.size_surface_standard")
add("native_delta_pp", round(n_nat - n_base, 1), "Native minus baseline standard share (pp)", "Sec 2.3 / Sec 4.6", "native_stress_sensitivity.json", "computed", "COMPUTED")
add("kt_gear", nss["K_t_gear"], "Gear root stress concentration Kt", "Sec 2.3", "native_stress_sensitivity.json", "K_t_gear")
add("kf_peterson", nss["K_f_peterson"], "FKM notch factor Kf (Peterson)", "Sec 2.3", "native_stress_sensitivity.json", "K_f_peterson")
add("kf_neuber", nss["K_f_neuber"], "FKM notch factor Kf (Neuber)", "Sec 2.3", "native_stress_sensitivity.json", "K_f_neuber")
add("agma_J_midpoint", nss["AGMA_J_range"]["J_midpoint"]["J"], "AGMA geometry factor J (midpoint)", "Sec 2.3", "native_stress_sensitivity.json", "AGMA_J_range.J_midpoint.J")
add("nbp_mean_bias_pp", nbp["mean_bias_pp"], "Native-chain mean bias of standard share (pp)", "Sec 4.6 / Constraints", "native_bias_propagation.json", "mean_bias_pp")
add("nbp_max_abs_bias_pp", nbp["max_abs_bias_pp"], "Native-chain max |bias| (pp)", "Sec 4.6 / Constraints", "native_bias_propagation.json", "max_abs_bias_pp")
add("nbp_native_mean_std", nbp["native_std_var_mean"], "Native-chain mean standard share (%)", "Sec 4.6 / Constraints", "native_bias_propagation.json", "native_std_var_mean")

# ---------- torque dependence ----------
add("t500_finite", torque["results"]["500"]["n_finite"], "Finite-life entries at 500 Nm", "Sec 3.7", "multi_torque_aft.json", "results.500.n_finite")
add("t500_censored", torque["results"]["500"]["n_censored"], "Censored entries at 500 Nm", "Sec 3.7", "multi_torque_aft.json", "results.500.n_censored")
add("t600_finite", torque["results"]["600"]["n_finite"], "Finite-life entries at 600 Nm", "Sec 3.7", "multi_torque_aft.json", "results.600.n_finite")
add("t600_censored", torque["results"]["600"]["n_censored"], "Censored entries at 600 Nm", "Sec 3.7", "multi_torque_aft.json", "results.600.n_censored")
add("t600_std_pct", torque["results"]["600"]["factors"]["size_surface_standard"], "Standard share at 600 Nm (%)", "Sec 3.7", "multi_torque_aft.json", "results.600.factors.size_surface_standard")
add("t600_rz_pct", torque["results"]["600"]["factors"]["rz_level"], "Roughness share at 600 Nm (%)", "Sec 3.7", "multi_torque_aft.json", "results.600.factors.rz_level")
add("t600_ms_pct", torque["results"]["600"]["factors"]["mean_stress_method"], "Mean-stress share at 600 Nm (%)", "Sec 3.7", "multi_torque_aft.json", "results.600.factors.mean_stress_method")

# ---------- Monte Carlo validation ----------
cov = [sim["simulation"]["bootstrap_coverage"][k] for k in sim["simulation"]["bootstrap_coverage"]]
max_abs_bias = max(abs(sim["simulation"]["per_factor"][k]["bias_pp"]) for k in sim["simulation"]["per_factor"])
add("mc_n_sims", sim["simulation"]["n_sims"], "Monte Carlo simulations", "Abstract / Sec 2.7 / Fig 2", "simulation_validation.json", "simulation.n_sims")
add("mc_max_abs_bias_pp", max_abs_bias, "Max |bias| of recovered shares (pp)", "Sec 2.7 / Fig 2", "simulation_validation.json", "simulation.per_factor.bias_pp", "COMPUTED")
add("mc_top3_recovery", round(sim["simulation"]["rank_recovery_top3_order"] * 100), "Top-3 rank recovery (%)", "Sec 2.7 / Fig 2", "simulation_validation.json", "simulation.rank_recovery_top3_order")
add("mc_R2", sim["sobol_equivalence"]["pooled_R2_LR_vs_Sobol"], "LR-vs-Sobol pooled R^2", "Sec 2.7 / Fig 2", "simulation_validation.json", "sobol_equivalence.pooled_R2_LR_vs_Sobol")
add("mc_cov_min", round(min(cov) * 100), "Bootstrap coverage minimum (%)", "Sec 2.7", "simulation_validation.json", "simulation.bootstrap_coverage", "COMPUTED")
add("mc_cov_max", round(max(cov) * 100), "Bootstrap coverage maximum (%)", "Sec 2.7", "simulation_validation.json", "simulation.bootstrap_coverage", "COMPUTED")
add("mc_bias_58pct", sim["censoring_envelope"]["0.58"]["bias_pp_dominant"], "Dominant-factor bias at 58% censoring (pp)", "Sec 2.7", "simulation_validation.json", "censoring_envelope.0.58.bias_pp_dominant")

# ---------- two-level variable amplitude ----------
add("tl_constant_spread", twolevel["summary"]["constant_700_Nm"]["spread_dex"], "Constant-amplitude spread (dex)", "Sec 4.5 / Fig 7", "two_level_variable_amplitude.json", "summary.constant_700_Nm.spread_dex")
add("tl_block_spread", twolevel["summary"]["two_level_700_800_Nm"]["spread_dex"], "Two-level block spread (dex)", "Sec 4.5 / Fig 7", "two_level_variable_amplitude.json", "summary.two_level_700_800_Nm.spread_dex")
add("tl_block_finite", twolevel["summary"]["two_level_700_800_Nm"]["n_finite"], "Finite entries under block loading", "Sec 4.5 / Fig 7", "two_level_variable_amplitude.json", "summary.two_level_700_800_Nm.n_finite")
add("tl_block_runout", twolevel["summary"]["two_level_700_800_Nm"]["n_runout"], "Run-outs under block loading", "Sec 4.5 / Fig 7", "two_level_variable_amplitude.json", "summary.two_level_700_800_Nm.n_runout")
for fac in ["mean_stress_method", "size_surface_standard", "sn_source", "damage_method"]:
    key = fac.replace(" ", "_")
    add("tl_block_" + key + "_pct", twolevel["two_level_shares_%"][fac], fac + " share under block loading (%)", "Sec 4.5 / Fig 7", "two_level_variable_amplitude.json", "two_level_shares_%." + fac)
add("tl_damage_diff", r(twolevel["summary"]["constant_700_Nm"]["damage_pair_abs_log10_diff"]["median"], 3), "Damage-rule pair |log10 diff|", "Sec 4.5", "two_level_variable_amplitude.json", "summary.constant_700_Nm.damage_pair_abs_log10_diff.median")

# ---------- Sobol censoring gradient ----------
for pt in gradient["points"]:
    key = str(int(pt["torque_Nm"]))
    add("sg_" + key + "_gap_pp", pt["gap_pp"], "Sobol-AFT gap at " + key + " Nm (pp)", "Sec 4.7 / Fig 10", "sobol_censoring_gradient.json", "points[" + key + "].gap_pp")
    add("sg_" + key + "_cens_pct", pt["censoring_%"], "Censoring at " + key + " Nm (%)", "Sec 4.7 / Fig 10", "sobol_censoring_gradient.json", "points[" + key + "].censoring_%")

# ---------- multi S-N sensitivity (Fig 8) ----------
add("msn_baseline_sn_pct", r(msn["summary"]["baseline_2_curves"]["sn_source_share_%"]), "S-N source share, 2 Waterloo curves (%)", "Sec 4.2 / Fig 8", "multi_sn_sensitivity.json", "summary.baseline_2_curves.sn_source_share_%")
add("msn_extended_sn_pct", r(msn["summary"]["extended_5_curves"]["sn_source_share_%"]), "S-N source share, 5 curves (%)", "Sec 4.2 / Fig 8", "multi_sn_sensitivity.json", "summary.extended_5_curves.sn_source_share_%")
add("msn_change_pp", msn["summary"]["change"]["sn_share_pp"], "S-N share change (pp)", "Sec 4.2 / Fig 8", "multi_sn_sensitivity.json", "summary.change.sn_share_pp")
sc = msn["summary"]["synthetic_curves"]
add("msn_s68_strength", r(sc["Waterloo_SMDIdbase_Iter068"]["sigma_w_1e6"]), "Waterloo #68 fatigue strength at 1e6 (MPa)", "Fig 8 caption / Sec 2.2", "multi_sn_sensitivity.json", "summary.synthetic_curves.Waterloo_SMDIdbase_Iter068.sigma_w_1e6")
add("msn_s66_strength", r(sc["Waterloo_SMDIdbase_Iter066"]["sigma_w_1e6"]), "Waterloo #66 fatigue strength at 1e6 (MPa)", "Fig 8 caption / Sec 2.2", "multi_sn_sensitivity.json", "summary.synthetic_curves.Waterloo_SMDIdbase_Iter066.sigma_w_1e6")
add("msn_high_strength", sc["Synthetic_42CrMo4_HighStrength"]["sigma_w_1e6"], "Synthetic high-strength 1e6 strength (MPa)", "Fig 8 caption", "multi_sn_sensitivity.json", "summary.synthetic_curves.Synthetic_42CrMo4_HighStrength.sigma_w_1e6")
add("msn_classic500", r(sc["Synthetic_42CrMo4_Classic500"]["sigma_w_1e6"]), "Synthetic classic-500 1e6 strength (MPa)", "Fig 8 caption", "multi_sn_sensitivity.json", "summary.synthetic_curves.Synthetic_42CrMo4_Classic500.sigma_w_1e6")
add("msn_classic560", sc["Synthetic_42CrMo4_Classic560"]["sigma_w_1e6"], "Synthetic classic-560 1e6 strength (MPa)", "Fig 8 caption", "multi_sn_sensitivity.json", "summary.synthetic_curves.Synthetic_42CrMo4_Classic560.sigma_w_1e6")

# ---------- K_A load factor ----------
add("ka10_censored", ka["results"]["1.0"]["n_censored"], "Censored entries at K_A=1.0", "Sec 3.5 / Sec 4.6", "ka_sensitivity.json", "results.1.0.n_censored")
add("ka125_censored", ka["results"]["1.25"]["n_censored"], "Censored entries at K_A=1.25", "Sec 3.5 / Sec 4.6", "ka_sensitivity.json", "results.1.25.n_censored")
add("ka15_censored", ka["results"]["1.5"]["n_censored"], "Censored entries at K_A=1.5", "Sec 3.5 / Sec 4.6", "ka_sensitivity.json", "results.1.5.n_censored")
add("ka125_sn_pct", ka["results"]["1.25"]["variance_fractions"]["sn_source"], "S-N share at K_A=1.25 (%)", "Sec 3.5", "ka_sensitivity.json", "results.1.25.variance_fractions.sn_source")
add("ka15_sn_pct", ka["results"]["1.5"]["variance_fractions"]["sn_source"], "S-N share at K_A=1.5 (%)", "Sec 3.5", "ka_sensitivity.json", "results.1.5.variance_fractions.sn_source")

# ---------- threshold sensitivity ----------
add("thr_1e6_cens_pct", threshold["fatigue_strength_criterion"]["1000000.0"]["censoring_rate_%"], "Censoring at 1e6 threshold (%)", "Sec 4.1", "threshold_sensitivity.json", "fatigue_strength_criterion.1000000.0.censoring_rate_%")
add("thr_3e6_cens_pct", threshold["basquin_only"]["bq_3000000.0"]["censoring_rate_%"], "Censoring at 3e6 threshold, Basquin-only (%)", "Sec 4.1", "threshold_sensitivity.json", "basquin_only.bq_3000000.0.censoring_rate_%")
add("thr_5e6_cens_pct", threshold["basquin_only"]["bq_5000000.0"]["censoring_rate_%"], "Censoring at 5e6 threshold, Basquin-only (%)", "Sec 4.1", "threshold_sensitivity.json", "basquin_only.bq_5000000.0.censoring_rate_%")
add("thr_none_cens_pct", threshold["basquin_only"]["bq_none"]["censoring_rate_%"], "Censoring with no threshold, Basquin-only (%)", "Sec 4.1", "threshold_sensitivity.json", "basquin_only.bq_none.censoring_rate_%")

# ---------- Wang 2023 benchmark ----------
add("wang_m", wang["wang_gear"]["m"], "Wang test gear module (mm)", "Sec 4.3", "benchmark_wang2023.json", "wang_gear.m")
add("wang_z", wang["wang_gear"]["z"], "Wang test gear teeth", "Sec 4.3", "benchmark_wang2023.json", "wang_gear.z")
add("wang_b", wang["wang_gear"]["b"], "Wang test gear face width (mm)", "Sec 4.3", "benchmark_wang2023.json", "wang_gear.b")
add("wang_stress_max", wang["wang_stress_levels_MPa"][0], "Wang highest stress level (MPa)", "Sec 4.3", "benchmark_wang2023.json", "wang_stress_levels_MPa[0]")
add("wang_stress_min", wang["wang_stress_levels_MPa"][-1], "Wang lowest stress level (MPa)", "Sec 4.3", "benchmark_wang2023.json", "wang_stress_levels_MPa[-1]")
add("wang_limit_max", wang["wang_fatigue_limit_MPa_at_1e7"][0], "Wang fatigue limit upper (MPa)", "Sec 4.3", "benchmark_wang2023.json", "wang_fatigue_limit_MPa_at_1e7[0]")
add("wang_limit_min", wang["wang_fatigue_limit_MPa_at_1e7"][-1], "Wang fatigue limit lower (MPa)", "Sec 4.3", "benchmark_wang2023.json", "wang_fatigue_limit_MPa_at_1e7[-1]")
add("wang_scatter_dex", wang["wang_reported_scatter_dex"], "Wang reported life scatter (dex)", "Sec 4.3", "benchmark_wang2023.json", "wang_reported_scatter_dex")

# ---------- literature anchors (manually verified against local PDFs) ----------
add("ulrich_P50_endurance", 624.5, "Ulrich et al. 2025 P50 endurance limit, unnotched R=-1 staircase (MPa)", "Sec 4.3 / Table 3", "paper/ref_papers/ulrich2025_42CrMo4_s10010-025-00818-x.pdf", "p.7 staircase result", "LIT")
add("ulrich_improvement_pct", "26-51", "Ulrich 2025 strength improvement vs conventional QT (%)", "Sec 4.3", "paper/ref_papers/ulrich2025_42CrMo4_s10010-025-00818-x.pdf", "p.7", "LIT")
add("ulrich_uts_range", "1322-1660", "Ulrich 2025 tempering-series UTS range (MPa)", "Sec 4.3 / Table 3", "paper/ref_papers/ulrich2025_42CrMo4_s10010-025-00818-x.pdf", "tempering series", "LIT")
add("w68_sigma_f", 1356.0, "Waterloo #68 Basquin sigma_f' (MPa)", "Sec 2.2 / Table 3", "multi_sn_sensitivity.json + ref_papers/Iter_068.pdf", "synthetic_curves.Waterloo_SMDIdbase_Iter068.sigma_f_prime", "LIT+JSON")
add("w68_b", sc["Waterloo_SMDIdbase_Iter068"]["b"], "Waterloo #68 Basquin exponent b", "Sec 2.2", "multi_sn_sensitivity.json + ref_papers/Iter_068.pdf", "synthetic_curves.Waterloo_SMDIdbase_Iter068.b", "LIT+JSON")
add("w66_sigma_f", 1684.0, "Waterloo #66 Basquin sigma_f' (MPa)", "Sec 2.2", "multi_sn_sensitivity.json + ref_papers/Iter_066.pdf", "synthetic_curves.Waterloo_SMDIdbase_Iter066.sigma_f_prime", "LIT+JSON")
add("w66_b", sc["Waterloo_SMDIdbase_Iter066"]["b"], "Waterloo #66 Basquin exponent b", "Sec 2.2", "multi_sn_sensitivity.json + ref_papers/Iter_066.pdf", "synthetic_curves.Waterloo_SMDIdbase_Iter066.b", "LIT+JSON")
add("w68_uts", sc["Waterloo_SMDIdbase_Iter068"]["uts"], "Waterloo #68 measured UTS (MPa)", "Sec 2.2 / Table 3", "multi_sn_sensitivity.json + ref_papers/Iter_068.pdf", "synthetic_curves.Waterloo_SMDIdbase_Iter068.uts", "LIT+JSON")
add("w66_uts", sc["Waterloo_SMDIdbase_Iter066"]["uts"], "Waterloo #66 measured UTS (MPa)", "Sec 2.2 / Table 3", "multi_sn_sensitivity.json + ref_papers/Iter_066.pdf", "synthetic_curves.Waterloo_SMDIdbase_Iter066.uts", "LIT+JSON")

# ---------- meta ----------
labels["_meta"] = {
    "generated_by": "generate_labels_v3.py",
    "date": "2026-08-06",
    "paper_version": "v4.8/v4.9",
    "n_labels": len(labels) - 1,
    "source_files": sorted({r[3] for r in rows}),
}

with open(os.path.join(OUT, "labels.json"), "w", encoding="utf-8") as f:
    json.dump(labels, f, indent=2, ensure_ascii=False)
print("Saved %d labels to output/labels.json" % (len(labels) - 1))

# ---------- ledger CSV + XLSX ----------
base = "\u6570\u5b57\u6765\u6e90-\u9f7f\u8f6e\u75b2\u52b3v4.8"
csv_path = os.path.join(HERE, base + ".csv")
xlsx_path = os.path.join(HERE, base + ".xlsx")
header = ["Paper Section", "Claim/Number", "Value in Paper", "Source File", "Source Key", "Match"]

with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
    import csv
    w = csv.writer(f)
    w.writerow(header)
    w.writerows(rows)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "numbers_v4.8"
    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    def _cellval(v):
        return v if isinstance(v, (int, float, str, bool)) or v is None else str(v)
    for row in rows:
        ws.append([_cellval(v) for v in row])
    widths = [28, 55, 16, 42, 48, 10]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.auto_filter.ref = "A1:F%d" % (len(rows) + 1)
    wb.save(xlsx_path)
    print("Saved %d ledger rows to:" % len(rows))
    print("  ", csv_path)
    print("  ", xlsx_path)
except Exception as e:
    print("XLSX write failed:", e)
